#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HanaXellOcr 대체 — 맥 Excel 변환 도구
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
parser_v7.py 의 OCR 파이프라인으로 PDF 등기부등본 → 6시트 Excel

[HanaXellOcr0.7 역공학 기반]
- output.xlsx 6시트 구조 → openpyxl로 재현
- 병합셀 규칙: B1:H1(문서헤더), C4:H4(기본정보), C6:H6(구분헤더) 등
- 참고: 역공학_백서.md §3.1-§3.3
"""

import os, sys, re, tkinter as tk
from tkinter import filedialog
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# parser_v7 import
PROJECT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(PROJECT_DIR))

from parser_v7 import (
    ocr_pdf, clean_text, parse_sections, detect_table_regions,
    validate_accuracy
)

# ============================================================
# 상수 (HanaXellOcr0.7 output.xlsx 기준)
# ============================================================
SHEET_NAMES = ['전체요약', '갑구', '을구', '공동담보목록', '요약-갑구', '요약-을구']

HEADER_FONT = Font(name='맑은 고딕', bold=True, size=14, color='1F4E79')
SECTION_FONT = Font(name='맑은 고딕', bold=True, size=12)
LABEL_FONT = Font(name='맑은 고딕', bold=True, size=10)
DATA_FONT = Font(name='맑은 고딕', size=10)
META_FONT = Font(name='맑은 고딕', italic=True, size=9, color='888888')

HEADER_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SECTION_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


# ============================================================
# 파일 선택
# ============================================================
def select_pdf():
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    path = filedialog.askopenfilename(
        title='등기부등본 PDF 선택',
        filetypes=[('PDF 파일', '*.pdf')]
    )
    root.destroy()
    return path if path else None


# ============================================================
# 파싱 파이프라인
# ============================================================
def parse_pdf(pdf_path):
    """PDF → OCR → 정제 → 섹션 → dict"""
    pdf_path = Path(pdf_path)
    pages = ocr_pdf(pdf_path)
    combined_raw = '\n\n'.join(p['text'] for p in pages)
    cleaned = clean_text(combined_raw)
    sections = parse_sections(cleaned)
    tables = detect_table_regions(pdf_path)
    return {
        'pdf_name': pdf_path.name,
        'pages': len(pages),
        'tables': tables,
        'sections': sections.get('sections', {}),
    }


# ============================================================
# Excel 렌더링 (openpyxl)
# ============================================================
def render_to_excel(data, output_path=None):
    """
    파싱 결과를 HanaXellOcr0.7 스타일 6시트 Excel로 렌더링
    
    [역공학 참고] output.xlsx 구조 (역공학_백서.md §3.1):
    - 전체요약(85x8), 갑구(15x6), 을구(22x6)
    - 공동담보목록(51x8), 요약-갑구(11x6), 요약-을구(14x6)
    - 병합셀: 문서헤더(B1:H1), 기본정보(C4:H4), 구분헤더(C6:H6)
    """
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    sections = data['sections']
    pdf_name = data['pdf_name']
    total_pages = data['pages']
    tables_count = len(data['tables'])
    
    _render_overview(wb, pdf_name, total_pages, tables_count, sections)
    _render_gapgu(wb, sections)
    _render_eulgu(wb, sections)
    _render_collateral(wb, sections)
    _render_summary(wb, sections, '갑구')
    _render_summary(wb, sections, '을구')
    
    # 시트 순서 정렬
    ordered = []
    for name in SHEET_NAMES:
        if name in wb.sheetnames:
            ordered.append(name)
    for i, name in enumerate(ordered):
        idx = wb.sheetnames.index(name)
        wb.move_sheet(name, offset=i - idx)
    
    if output_path:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path
    return wb


def _render_overview(wb, pdf_name, total_pages, tables_count, sections):
    """[시트1] 전체요약"""
    ws = wb.create_sheet('전체요약')
    r = 1
    
    # === Row 1: 문서 헤더 [역공학: B1:H1 병합] ===
    _merge_and_set(ws, r, 1, r, 8, '등기사항전부증명서 (발췌 자동화)',
                   HEADER_FONT, CENTER)
    r += 2
    
    # 문서 메타정보
    meta_lines = [
        f'원본 PDF: {pdf_name}    |    페이지: {total_pages}페이지    |    표 영역: {tables_count}개',
        f'HanaXellOcr0.7 역공학 기반 — 맥 Excel 확장앱 v1.0  (parser_v7 + openpyxl)',
    ]
    for line in meta_lines:
        _merge_and_set(ws, r, 1, r, 8, line, META_FONT)
        r += 1
    r += 1
    
    # === 기본정보 섹션 [역공학: C4:H4] ===
    _merge_and_set(ws, r, 1, r, 8, '기  본  정  보', SECTION_FONT, CENTER)
    _set_fill(ws, r, 1, r, 8, SECTION_FILL)
    r += 1
    
    info = _extract_basic_info(sections)
    for label, value in [
        ('고유번호', info.get('고유번호', '')),
        ('소재지', info.get('소재지', '')),
        ('부동산종류', info.get('부동산종류', '')),
        ('열람일시', info.get('열람일시', '')),
    ]:
        _set_cell(ws, r, 1, label, LABEL_FONT)
        _merge_and_set(ws, r, 2, r, 4, value if value else '(추출 실패)', DATA_FONT)
        r += 1
    
    r += 1
    
    # === 표제부 내용 ===
    _write_section_block(ws, r, sections, '【표제부】')
    
    _set_col_widths(ws, [15, 22, 18, 22, 40, 15, 18, 18])


def _render_gapgu(wb, sections):
    """[시트2] 갑구"""
    ws = wb.create_sheet('갑구')
    _write_document_header(ws, '갑구 — 소유권에 관한 사항')
    _write_section_block(ws, 5, sections, '【갑구】')
    _set_col_widths(ws, [12, 18, 16, 20, 40, 15])


def _render_eulgu(wb, sections):
    """[시트3] 을구"""
    ws = wb.create_sheet('을구')
    _write_document_header(ws, '을구 — 소유권 외의 권리에 관한 사항')
    _write_section_block(ws, 5, sections, '【을구】')
    _set_col_widths(ws, [12, 18, 16, 20, 40, 15])


def _render_collateral(wb, sections):
    """[시트4] 공동담보목록"""
    ws = wb.create_sheet('공동담보목록')
    _write_document_header(ws, '공동담보목록')
    _write_section_block(ws, 5, sections, '【공동담보목록】')
    _set_col_widths(ws, [15, 22, 18, 22, 40, 15, 18, 18])


def _render_summary(wb, sections, section_type):
    """[시트5/6] 요약-갑구 / 요약-을구"""
    ws = wb.create_sheet(f'요약-{section_type}')
    target = f'【{section_type}】'
    r = 1
    
    _merge_and_set(ws, r, 1, r, 6, f'요약 — {section_type}', HEADER_FONT, CENTER)
    r += 2
    
    # 헤더
    for c, h in enumerate(['순위번호', '등기목적', '권리자', '금액/지분', '비고'], 1):
        _set_cell(ws, r, c, h, LABEL_FONT, HEADER_FILL, THIN_BORDER, CENTER)
    r += 1
    
    # 데이터
    if target in sections:
        for line in sections[target]:
            line = line.strip()
            if not line or line.startswith('【') or line.startswith('['):
                continue
            _merge_and_set(ws, r, 1, r, 5, line, DATA_FONT)
            _apply_border(ws, r, 1, r, 5)
            r += 1
    
    _set_col_widths(ws, [12, 18, 25, 20, 25])


def _write_document_header(ws, title):
    """공통 문서 헤더"""
    _merge_and_set(ws, 1, 1, 1, 8, title, HEADER_FONT, CENTER)


def _write_section_block(ws, start_row, sections, section_name):
    """섹션 이름이 있으면 내용을 기록"""
    r = start_row
    if section_name not in sections:
        _set_cell(ws, r, 1, f'[{section_name} — 데이터 없음]', META_FONT)
        return r + 1
    
    lines = sections[section_name]
    
    # 섹션 헤더
    _merge_and_set(ws, r, 1, r, 8, section_name, SECTION_FONT, CENTER)
    _set_fill(ws, r, 1, r, 8, SECTION_FILL)
    r += 1
    
    # 데이터 행
    for line in lines:
        line = line.strip()
        if not line:
            continue
        if line.startswith('【') or line.startswith('['):
            continue
        if any(h in line for h in ['순위번호', '등기목적', '접수']):
            continue
        
        _merge_and_set(ws, r, 1, r, 6, line, DATA_FONT)
        _apply_border(ws, r, 1, r, 6)
        r += 1
    
    return r


# ============================================================
# openpyxl 헬퍼
# ============================================================
def _merge_and_set(ws, r1, c1, r2, c2, value, font=None, alignment=None):
    if r1 != r2 or c1 != c2:
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(row=r1, column=c1)
    cell.value = value
    if font: cell.font = font
    if alignment: cell.alignment = alignment

def _set_cell(ws, r, c, value, font=None, fill=None, border=None, alignment=None):
    cell = ws.cell(row=r, column=c)
    cell.value = value
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if alignment: cell.alignment = alignment

def _set_fill(ws, r1, c1, r2, c2, fill):
    for row in range(r1, r2+1):
        for col in range(c1, c2+1):
            ws.cell(row=row, column=col).fill = fill

def _apply_border(ws, r1, c1, r2, c2):
    for row in range(r1, r2+1):
        for col in range(c1, c2+1):
            ws.cell(row=row, column=col).border = THIN_BORDER

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _extract_basic_info(sections):
    """섹션에서 기본정보 필드 추출"""
    info = {}
    all_lines = []
    for lines in sections.values():
        all_lines.extend(lines)
    full_text = '\n'.join(all_lines)
    
    patterns = {
        '고유번호': r'고유번호\s*[:\|]?\s*(\S+)',
        '소재지': r'소재지\s*[:\|]?\s*(.+?)(?=부동산종류|열람일시|$)',
        '부동산종류': r'부동산종류\s*[:\|]?\s*(\S+)',
        '열람일시': r'열람일시\s*[:\|]?\s*(.+?)(?:\n|$)',
    }
    for key, pat in patterns.items():
        m = re.search(pat, full_text)
        if m:
            info[key] = m.group(1).strip()
    return info


# ============================================================
# 메인 진입점
# ============================================================
def run_pipeline(pdf_path=None, output_dir=None):
    """전체 파이프라인 실행 (GUI/CLI 공통)"""
    if pdf_path is None:
        pdf_path = select_pdf()
    if not pdf_path:
        return None
    
    print(f'[등기부 파서] {Path(pdf_path).name} 파싱 시작...')
    
    # 1) 파싱
    data = parse_pdf(pdf_path)
    print(f'  OCR: {data["pages"]}페이지 완료')
    print(f'  표 영역: {len(data["tables"])}개')
    print(f'  섹션: {list(data["sections"].keys())}')
    
    # 2) 출력
    if output_dir is None:
        output_dir = PROJECT_DIR / 'experiments'
    output_dir = Path(output_dir)
    output_path = output_dir / f'addin_{Path(pdf_path).stem}.xlsx'
    
    result = render_to_excel(data, output_path)
    print(f'  저장: {result}')
    return result


if __name__ == '__main__':
    import argparse
    ap = argparse.ArgumentParser(description='등기부등본 PDF → Excel 변환')
    ap.add_argument('pdf', nargs='?', help='PDF 파일 경로')
    args = ap.parse_args()
    
    pdf = args.pdf
    if not pdf:
        pdf = PROJECT_DIR / '2849-2018-019318_25696174641_RIS.pdf'
        if not pdf.exists():
            print('PDF 파일을 지정해주세요.')
            sys.exit(1)
    
    result = run_pipeline(str(pdf))
    if result:
        print(f'\n✅ 완료! {result}')
    else:
        print('\n❌ 실패')
