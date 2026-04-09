"""
등기부등본 → Excel v6 (v1.0.2)
수정:
  1. 매매목록/공동담보목록: 목록번호를 각 데이터행 B열(첫 컬럼)에 표시
     + 목록번호 구분 헤더행 추가
  2. 요약_소유지분: 컬럼 분리 정확히 (주민번호/최종지분/주소 각 셀)
  3. 요약_갑구/을구: 순위번호가 등기목적에 붙지 않도록 컬럼 정리
"""
import os, csv
from typing import Dict, List
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KO = "맑은 고딕"
SZ = 11

BG_TITLE  = "D9D9D9"
BG_COLHDR = "EFEFEF"
BG_ALT    = "F7F7F7"
BG_EVEN   = "FFFFFF"
BD_COLOR  = "C0C0C0"
BG_INFO   = "E8E8E8"
BG_LISTNO = "EFEFEF"   # 목록번호 구분행 배경

def _s(st="thin", co=BD_COLOR): return Side(style=st, color=co)
def _bd(): return Border(left=_s(),right=_s(),top=_s(),bottom=_s())
def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, sz=SZ, color="000000"): return Font(name=KO,bold=bold,size=sz,color=color)
def _al(h="left",v="top",wrap=True): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

COL_OFFSET = 1  # A열 여백

# 섹션별 컬럼 (헤더명, 너비)
COLS = {
    "기본정보":      [("항목",14),("내용",76)],
    "표제부":        [("표시번호",7),("접수",16),("소재지번",28),("지목",6),("면적",10),("등기원인및기타사항",40)],
    "갑구":          [("순위번호",8),("등기목적",18),("접수",18),("등기원인",18),("권리자및기타사항",46)],
    "을구":          [("순위번호",8),("등기목적",18),("접수",18),("등기원인",18),("권리자및기타사항",46)],
    # 매매목록: 목록번호 첫 열, 나머지
    "매매목록":      [("목록번호",11),("일련번호",7),("부동산표시",36),("순위번호",8),("등기원인",18),("경정원인",14)],
    # 공동담보목록: 목록번호 첫 열
    "공동담보목록":  [("목록번호",11),("일련번호",7),("부동산표시",28),("관할등기소",24),("순위번호",8),("생성원인",22),("변경소멸",22)],
    "요약_소유지분": [("등기명의인",18),("주민등록번호",16),("최종지분",10),("주소",34),("순위번호",8)],
    "요약_갑구":     [("순위번호",8),("등기목적",16),("접수정보",20),("주요등기사항",36),("대상소유자",12)],
    "요약_을구":     [("순위번호",8),("등기목적",16),("접수정보",20),("주요등기사항",36),("대상소유자",12)],
}

LABEL = {
    "기본정보":      "기  본  정  보",
    "표제부":        "【 표 제 부 】  ( 토지의 표시 )",
    "갑구":          "【 갑 구 】  ( 소유권에 관한 사항 )",
    "을구":          "【 을 구 】  ( 소유권 이외의 권리에 관한 사항 )",
    "매매목록":      "【 매 매 목 록 】",
    "공동담보목록":  "【 공 동 담 보 목 록 】",
    "요약_소유지분": "주요 등기사항 요약  ▷  1. 소유지분현황 ( 갑구 )",
    "요약_갑구":     "주요 등기사항 요약  ▷  2. 소유지분을 제외한 소유권에 관한 사항 ( 갑구 )",
    "요약_을구":     "주요 등기사항 요약  ▷  3. (근)저당권 및 전세권 등 ( 을구 )",
}

INFO_ROWS = [
    ("고유번호","고유번호"),("소재지","소재지"),
    ("부동산종류","부동산종류"),("열람일시","열람일시"),("현황","현황(요약)"),
]

ORDER = ["기본정보","표제부","갑구","을구",
         "매매목록","공동담보목록",
         "요약_소유지분","요약_갑구","요약_을구"]


def _h(values, widths):
    lines = 1
    for val, w in zip(values, widths):
        if not val: continue
        t = str(val)
        n = max(1, len(t) // max(int(w*1.6), 1) + t.count("\n") + 1)
        lines = max(lines, min(n, 12))
    return max(int(SZ*1.9), int(lines*SZ*1.7))

def _col(ci): return ci + COL_OFFSET

def _set_widths(ws, col_defs):
    ws.column_dimensions["A"].width = 2
    for ci, (_, w) in enumerate(col_defs, 1):
        ws.column_dimensions[get_column_letter(_col(ci))].width = w

def _cell(ws, row, col, val="", bold=False, bg=BG_EVEN, wrap=True, h="left", v="top"):
    c = ws.cell(row=row, column=col, value=val if val else "")
    c.font = _font(bold=bold, sz=SZ)
    c.fill = _fill(bg)
    c.alignment = _al(h, v, wrap)
    c.border = _bd()
    return c

def _write_title(ws, row, n_cols, text):
    for ci in range(1, n_cols+COL_OFFSET+1):
        ws.cell(row=row,column=ci).fill = _fill(BG_TITLE)
    c = ws.cell(row=row, column=_col(1), value=text)
    c.font = _font(bold=True, sz=SZ)
    c.alignment = _al("left","center",wrap=False)
    ws.row_dimensions[row].height = int(SZ*2.4)
    return row+1

def _write_col_hdr(ws, row, col_defs):
    for ci,(name,_) in enumerate(col_defs,1):
        c = _cell(ws, row, _col(ci), name, bold=True, bg=BG_COLHDR, h="center", v="center")
    ws.row_dimensions[row].height = int(SZ*2.0)
    return row+1

def _write_info(ws, row, n_cols, info):
    row = _write_title(ws, row, n_cols, LABEL["기본정보"])
    for field,label in INFO_ROWS:
        val = info.get(field,"")
        if not val: continue
        _cell(ws, row, _col(1), label, bold=True, bg=BG_INFO, h="center", v="center", wrap=False)
        _cell(ws, row, _col(2), val, bg=BG_EVEN, h="left", v="center", wrap=False)
        if n_cols > 2:
            ws.merge_cells(start_row=row,start_column=_col(2),
                           end_row=row,end_column=_col(n_cols))
        ws.row_dimensions[row].height = int(SZ*1.9)
        row += 1
    return row+1


def _write_listno_block(ws, row, col_defs, listno_label, records, ri_offset=0):
    """목록번호 구분 헤더 + 해당 목록의 데이터 행 출력"""
    n = len(col_defs)
    col_names  = [c for c,_ in col_defs]
    col_widths = [w for _,w in col_defs]

    # 목록번호 구분 행
    for ci in range(1, n+1):
        bg = BG_LISTNO
        val = "목록번호" if ci==1 else (listno_label if ci==2 else "")
        bold = (ci <= 2)
        _cell(ws, row, _col(ci), val, bold=bold, bg=bg, h="left", v="center", wrap=False)
    ws.row_dimensions[row].height = int(SZ*1.8)
    row += 1

    # 컬럼 헤더
    row = _write_col_hdr(ws, row, col_defs)

    # 데이터 행
    for ri, rec in enumerate(records):
        if "내용" in rec and "부동산표시" not in rec:
            # 거래가액 등 단순 행
            for ci in range(1, n+1):
                val = rec.get("내용","") if ci==1 else ""
                _cell(ws, row, _col(ci), val, bg=BG_EVEN)
            ws.row_dimensions[row].height = int(SZ*1.9)
            row += 1
            continue

        values = [str(rec.get(cn,"")) for cn in col_names]
        ht = _h(values, col_widths)
        bg = BG_ALT if (ri+ri_offset)%2 else BG_EVEN
        for ci, val in enumerate(values, 1):
            _cell(ws, row, _col(ci), val, bg=bg, wrap=True)
        ws.row_dimensions[row].height = ht
        row += 1

    return row


def _write_data_grouped(ws, row, col_defs, records, sec_name):
    """매매목록/공동담보목록: 목록번호 단위로 그룹핑하여 출력"""
    from collections import OrderedDict
    groups = OrderedDict()
    for rec in records:
        listno = rec.get("목록번호","")
        if "내용" in rec and "부동산표시" not in rec:
            # 거래가액 행은 목록번호 없는 그룹에
            groups.setdefault(listno, []).append(rec)
        else:
            groups.setdefault(listno, []).append(rec)

    ri_global = 0
    for listno, recs in groups.items():
        row = _write_listno_block(ws, row, col_defs, listno, recs, ri_offset=ri_global)
        ri_global += len(recs)
        row += 1  # 목록번호 그룹 간 빈 행

    return row


def _write_data_normal(ws, row, col_defs, records):
    """일반 섹션 데이터 행 출력"""
    col_names  = [c for c,_ in col_defs]
    col_widths = [w for _,w in col_defs]
    n = len(col_defs)

    for ri, rec in enumerate(records):
        if "내용" in rec and "부동산표시" not in rec:
            for ci in range(1, n+1):
                val = rec.get("내용","") if ci==1 else ""
                _cell(ws, row, _col(ci), val, bg=BG_EVEN if ri%2==0 else BG_ALT)
            ws.row_dimensions[row].height = int(SZ*1.9)
            row += 1
            continue

        values = [str(rec.get(cn,"")) for cn in col_names]
        ht = _h(values, col_widths)
        bg = BG_ALT if ri%2 else BG_EVEN
        for ci, val in enumerate(values, 1):
            _cell(ws, row, _col(ci), val, bg=bg, wrap=True)
        ws.row_dimensions[row].height = ht
        row += 1

    return row+1


def build_sheet(ws, data, secs):
    ws.sheet_view.showGridLines = False
    n_cols = max((len(COLS[s]) for s in secs if s in COLS and data.get(s)), default=5)

    # 상단 타이틀
    meta = data.get("_meta",[{}])[0]
    ts   = datetime.now().strftime("%Y-%m-%d %H:%M")
    title = (f"등기사항전부증명서  |  {meta.get('소재지','')}  |  "
             f"고유번호 {meta.get('고유번호','')}  |  {ts}")
    for ci in range(1, n_cols+COL_OFFSET+1):
        ws.cell(row=1,column=ci).fill = _fill(BG_COLHDR)
    c = ws.cell(row=1, column=_col(1), value=title)
    c.font = _font(bold=False, sz=SZ, color="444444")
    c.alignment = _al("left","center",wrap=False)
    ws.merge_cells(start_row=1,start_column=_col(1),end_row=1,end_column=_col(n_cols))
    ws.row_dimensions[1].height = int(SZ*2.2)
    cur = 3

    for sec in secs:
        rows = data.get(sec,[])
        if not rows: continue
        col_defs = COLS.get(sec)
        if not col_defs: continue

        if sec == "기본정보":
            cur = _write_info(ws, cur, n_cols, rows[0])
            continue

        cur = _write_title(ws, cur, n_cols, LABEL.get(sec,sec))

        if sec in ("매매목록","공동담보목록"):
            # 목록번호 단위 그룹 출력 (헤더 포함)
            cur = _write_data_grouped(ws, cur, col_defs, rows, sec)
        else:
            cur = _write_col_hdr(ws, cur, col_defs)
            cur = _write_data_normal(ws, cur, col_defs, rows)

    # 열 너비
    ref = COLS.get("을구",[])
    if ref: _set_widths(ws, ref)
    for sec in secs:
        if sec not in ("기본정보",) and sec in COLS and data.get(sec):
            _set_widths(ws, COLS[sec])
            break


def export_xlsx(data: Dict[str,list], out_path: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet("전체요약")
    build_sheet(ws_all, data, ORDER)
    _set_widths(ws_all, COLS["을구"])
    ws_all.column_dimensions["A"].width = 2

    for sec in ORDER:
        if sec == "기본정보": continue
        if not data.get(sec): continue
        if sec not in COLS: continue
        ws = wb.create_sheet(sec.replace("요약_","요약-"))
        build_sheet(ws, data, ["기본정보",sec])
        _set_widths(ws, COLS[sec])
        ws.column_dimensions["A"].width = 2

    wb.save(out_path)
    return out_path


def export_csv(data: Dict[str,list], out_dir: str) -> List[str]:
    os.makedirs(out_dir, exist_ok=True)
    paths = []
    for sec in ORDER:
        rows = data.get(sec,[])
        if not rows: continue
        keys = []
        for r in rows:
            for k in r:
                if k not in keys: keys.append(k)
        fname = os.path.join(out_dir, f"{sec}.csv")
        with open(fname,"w",newline="",encoding="utf-8-sig") as f:
            w = csv.DictWriter(f,fieldnames=keys,extrasaction="ignore")
            w.writeheader(); w.writerows(rows)
        paths.append(fname)
    return paths
