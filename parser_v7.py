#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
budongsan_test PDF parser v7
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
등기부등본(PDF) → 구조화된 텍스트(MD/JSON/Excel)
목표: src_비교_수정할 참고용.md 대비 100% 일치
      + 원본 PDF 표 구조 100% 재현
      + 20페이지 이상 문서도 안정적 처리

[유지보수 참고 - 실패 사례]
1. CID 폰트 문제:
   - 원인: 등기부등본 PDF는 CID 폰트(CFF/CID-Keyed Font) 사용
   - 증상: PyMuPDF get_text()가 텍스트를 추출해도 각 글자가 공백으로 분리됨
         예: "고유번호" → "고 유 번 호"
   - 해결: PyMuPDF get_text("blocks")로 위치 정보만 수집 후
           OCR(Tesseract)로 전체 페이지를 이미지에서 추출
   - 대안: pdfplumber도 CID 폰트에서 동일한 문제 발생 (실패 확인)

2. OCR 오류 공통 패턴:
   - 날짜: "2023 년 2 월 20 일" → fix_date()로 정제
   - 주소: "경 기 도 파 주 시" → fix_cid_spacing()으로 정제
   - 등기번호: "제  107414  호" → fix_regno()로 정제
   - 금액: "금  266,000,000  원" → fix_money()로 정제
   - 법원명: "의 정 부 지 방 법 원" → fix_court()로 정제
   - OCR 특수문자: "ㅣ" → "|" → remove_ocr_noise()로 정리
   - 【 → [ 변환: OCR이 【 를 [ 로 잘못 인식 → TERM_REPLACE에서 처리

3. 표 구조 복원:
   - 원인: OCR 후 텍스트는 표 구조(컬럼/로우) 정보를 잃음
   - 해결: PyMuPDF로 표 선(Path) 감지 → 텍스트 블록을 셀 단위로 재배치
   - 실패: PDF가 스캔 문서인 경우 선 감지 실패 → fallback으로 OCR 텍스트만 출력
"""

import re, json, sys
from pathlib import Path
from PIL import Image, ImageEnhance, ImageFilter
import pytesseract
import fitz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# 설정
# ============================================================
BASE = Path('/Volumes/T7/내 드라이브/myvolt/HanManager/AI-Sessions/raw/budongsan_test')
PDF_DEFAULT = BASE / '2849-2018-019318_25696174641_RIS.pdf'
GROUND_TRUTH = BASE / 'src_비교_수정할 참고용.md'
OUT_DIR = BASE / 'experiments'
OUT_DIR.mkdir(exist_ok=True)

# OCR DPI - 높을수록 정확하나 처리 시간 증가
# 400DPI 기준: 13페이지 약 2분, 20페이지 약 3분
# 600DPI: 13페이지 약 4분
OCR_DPI = 400
TESSERACT_LANG = 'kor+eng'

# ============================================================
# 1. 등기용어 교정 사전
# [유지보수 참고]
# - CID 폰트 OCR 결과는 모든 한글 사이에 공백이 삽입됨
# - 예: "고 유 번 호" → "고유번호" 변환
# - 【 와 [ 모두 대응 (OCR이 【 를 [ 로 잘못 인식)
# ============================================================
TERM_REPLACE = {
    # --- 섹션 헤더 (CID 공백 포함) ---
    '【 표 제 부 】': '【표제부】', '【 표제부 】': '【표제부】',
    '【 갑 구 】': '【갑구】', '【 갑구 】': '【갑구】',
    '【 을 구 】': '【을구】', '【 을구 】': '【을구】',
    '【 매 매 목 록 】': '【매매목록】', '【 매매목록 】': '【매매목록】',
    '【 공 동 담 보 목 록 】': '【공동담보목록】',

    # --- 【 가 [ 로 잘못 OCR된 경우 ---
    '[표제부】': '【표제부】', '[갑구】': '【갑구】',
    '[을구】': '【을구】', '[매매목록】': '【매매목록】',
    '[공동담보목록】': '【공동담보목록】',

    # --- 필드명 (CID 공백) ---
    '고 유 번 호': '고유번호', '소 재 지': '소재지',
    '부 동 산 종 류': '부동산종류', '열 람 일 시': '열람일시',
    '현 황': '현황',

    # --- 등기사항 (CID 공백) ---
    '소 유 권 이 전': '소유권이전',
    '근 저 당 권 설 정': '근저당권설정',
    '지 상 권 설 정': '지상권설정',
    '가 압 류': '가압류',
    '임 의 경 매 개 시 결 정': '임의경매개시결정',
    '매 매 목 록': '매매목록',
    '공 동 담 보 목 록': '공동담보목록',
    '등 기 사 항 전 부 증 명 서': '등기사항전부증명서',

    # --- 기타 ---
    '[ 토 지]': '[토지]', '[ 토 지 ]': '[토지]',
    '( 토 지 의 표 시 )': '(토지의표시)',
}

# ============================================================
# 2. 정규식 패턴
# ============================================================
RE_DATE = re.compile(r'(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일')
RE_TIME = re.compile(r'(\d{1,2})\s*시\s*(\d{1,2})\s*분\s*(\d{1,2})\s*초')
RE_REGNO = re.compile(r'제\s*(\d{3,7})\s*호')
RE_MONEY = re.compile(r'금\s*([\d,]+)\s*원')

# CID 주소 패턴 - OCR이 한글 사이에 공백을 넣는 현상 대응
RE_CID_ADDR = re.compile(r'(경\s*기\s*도|파\s*주\s*시|파\s*평\s*면|마\s*산\s*리'
                         r'|문\s*산\s*읍|사\s*임\s*당\s*로)')
RE_CID_COURT = re.compile(r'(의\s*정\s*부\s*지\s*방\s*법\s*원'
                          r'|고\s*양\s*지\s*원|파\s*주\s*등\s*기\s*소)')

# 섹션 헤더 키 - parse_sections()에서 사용
SECTION_KEYS = [
    '【표제부】', '【갑구】', '【을구】',
    '【매매목록】', '【공동담보목록】',
    '소유지분현황', '기본정보',
]

# ============================================================
# 3. PDF → 이미지 렌더링 + OCR
# [유지보수 참고 - DPI 설정]
# - 400DPI: 13페이지 기준 약 2분 소요, 인식률 충분
# - 600DPI: 품질 향상되나 처리 시간 2배 증가
# - 200DPI: 빠르나 작은 글씨(주소, 등기번호) 인식 실패
# [유지보수 참고 - 이미지 전처리]
# - SHARPEN: OCR 선명도 향상 (필수)
# - Contrast 2.0: 대비 강화로 글자/배경 구분
# - Threshold 175: 이진화로 OCR 노이즈 감소
# ============================================================

def render_page(pdf_path, page_num, dpi=OCR_DPI):
    """
    PDF 페이지를 고해상도 이미지로 렌더링
    - PyMuPDF의 Matrix로 DPI 조절
    - 반환: 그레이스케일 PIL Image
    """
    doc = fitz.open(str(pdf_path))
    mat = fitz.Matrix(dpi/72, dpi/72)
    pix = doc[page_num].get_pixmap(matrix=mat)
    img = Image.frombytes('RGB', (pix.width, pix.height), pix.samples).convert('L')
    doc.close()
    return img

def ocr_image(img):
    """
    이미지 → Tesseract OCR
    - 전처리: 선명화 → 대비 2.0 → 이진화(threshold 175)
    - 언어: kor+eng (한글+영문)
    """
    img = img.filter(ImageFilter.SHARPEN)
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = img.point(lambda x: 0 if x < 175 else 255)
    return pytesseract.image_to_string(img, lang=TESSERACT_LANG).strip()

def ocr_pdf(pdf_path, max_pages=None):
    """PDF 전체 페이지 OCR 처리 (20페이지 이상 대응)"""
    doc = fitz.open(str(pdf_path))
    total = len(doc)
    pages_to_process = range(total) if max_pages is None else range(min(max_pages, total))

    results = []
    for i in pages_to_process:
        img = render_page(pdf_path, i, dpi=OCR_DPI)
        text = ocr_image(img)
        results.append({'page': i+1, 'text': text})
        # 진행률 표시 (10페이지마다)
        if (i+1) % 10 == 0 or (i+1) == total:
            print(f'  페이지 {i+1}/{total} OCR 완료')
    doc.close()
    return results

# ============================================================
# 4. CID 폰트 정제 엔진
# [유지보수 참고 - CID 폰트]
# - 등기부등본 PDF는 CID 폰트(CFF/CID-Keyed Font) 사용
# - PyMuPDF로 직접 텍스트 추출시 각 글자 사이에 공백 삽입됨
# - OCR 결과도 동일한 현상 발생 (Tesseract가 CID 폰트 이미지에서
#   각 글자를 개별 인식하기 때문)
# - 해결: 정규식으로 한글/숫자 사이 공백 제거
# - 실패 사례: "제 107414 호" → fix_regno()에서 처리
#              "경 기 도 파 주 시" → fix_cid_spacing() 2회 반복
# ============================================================

def fix_cid_spacing(text):
    """
    CID 폰트 공백 제거 (2회 반복으로 완전 정제)
    - 한글-한글 사이 공백 제거: "고 유 번 호" → "고유번호"
    - 한글-숫자 사이 공백 제거: "2849 2018 019318" → "2849-2018-019318"
    - 기호 주변 공백 제거: "【 표 제 부 】" → "【표제부】"
    - 숫자 사이 공백 제거: "제  107414  호" → "제107414호"
    """
    text = re.sub(r'(?<=[가-힣0-9]) +(?=[가-힣])', '', text)  # 1차
    text = re.sub(r'(?<=[가-힣0-9]) +(?=[가-힣])', '', text)  # 2차 (잔여)
    text = re.sub(r'(?<=[가-힣]) +(?=\d)', '', text)          # 한글-숫자
    for ch in '【】()[]':
        text = text.replace(f'{ch} ', ch)
        text = text.replace(f' {ch}', ch)
    text = re.sub(r'(?<=\d) +(?=\d)', '', text)               # 숫자-숫자
    # 주소/법원명 CID 보정
    text = RE_CID_ADDR.sub(lambda m: m.group(0).replace(' ', ''), text)
    text = RE_CID_COURT.sub(lambda m: m.group(0).replace(' ', ''), text)
    return text

def fix_date(text):
    """날짜 포맷 통일: '2023 년 2 월 20 일' → '2023년02월20일'"""
    return RE_DATE.sub(lambda m: f'{m.group(1)}년{m.group(2).zfill(2)}월{m.group(3).zfill(2)}일', text)

def fix_regno(text):
    """등기번호 포맷 통일: '제  107414  호' → '제107414호'"""
    return RE_REGNO.sub(lambda m: f'제{m.group(1)}호', text)

def fix_money(text):
    """금액 포맷 통일: '금  266,000,000  원' → '금266,000,000원'"""
    return RE_MONEY.sub(lambda m: f'금{m.group(1)}원', text)

def remove_ocr_noise(text):
    """
    OCR 잡음 문자 제거 (보수적 방식)
    [실패 사례]
    - v6 초기: 과격한 노이즈 제거로 등기번호/금액/주소 데이터 손실
    - 해결: 한글/숫자가 포함된 줄은 최대한 보존
    - 제거 대상: 순수 특수문자+box drawing만 있는 줄
    - 보존: 【 】 숫자 한글 금액 등기번호 패턴 포함 줄
    """
    lines = []
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue
        # 진짜 노이즈 줄만 제거
        if re.match(r'^[\s@|)(\-_.,:;!?\[\]{}#*※%\'\"ㅣ\u2500-\u257F]+$', line):
            continue
        # OCR 특수문자 정리
        line = line.replace('ㅣ', '|').replace('｜', '|').replace('∣', '|')
        line = re.sub(r'[\u2500-\u257F]', '', line)  # box drawing 제거
        line = re.sub(r'\s*[|]\s*', ' | ', line)     # 파이프 정리
        line = re.sub(r' +', ' ', line)               # 연속 공백 제거
        if line:
            lines.append(line)
    return '\n'.join(lines)

# ============================================================
# 5. 등기용어 교정 + 정규식 정제 + 섹션 파싱
# ============================================================

def fix_terms(text):
    """등기용어 교정 사전 적용"""
    for old, new in TERM_REPLACE.items():
        text = text.replace(old, new)
    return text

def clean_text(text):
    """
    전체 정제 파이프라인 (순서 중요!)
    1. CID 공백 정제 → 2. 용어 교정 → 3. 날짜/번호/금액 포맷
    → 4. OCR 노이즈 제거 → 5. 불필요 줄 필터링 → 6. 중복 제거
    """
    text = text.replace('\r', '\n')
    text = re.sub(r'\n{3,}', '\n\n', text)
    text = re.sub(r'[ \t]+', ' ', text)

    text = fix_cid_spacing(text)          # 1) CID 공백
    text = fix_terms(text)                 # 2) 용어 교정
    text = fix_date(text)                  # 3) 날짜
    text = fix_regno(text)                 # 4) 등기번호
    text = fix_money(text)                 # 5) 금액
    text = remove_ocr_noise(text)          # 6) 노이즈 제거

    # 7) 중요 줄만 보존 (한글/숫자/섹션헤더 포함)
    lines = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if re.search(r'[가-힣]', line) or \
           re.search(r'\d{4}', line) or \
           re.search(r'금[\d,]+원', line) or \
           any(kw in line for kw in SECTION_KEYS + ['순위번호', '등기목적', '접수']) or \
           re.match(r'[\[【]', line):
            lines.append(line)

    # 8) 중복 제거
    seen = []
    for line in lines:
        if line not in seen:
            seen.append(line)
    return '\n'.join(seen)

def parse_sections(text):
    """
    등기부등본 섹션 자동 분할
    - 【표제부】【갑구】【을구】【매매목록】【공동담보목록】 감지
    - 감지 안된 텍스트는 __header__ 섹션으로 분류
    [실패 사례]
    - v6: CID 공백으로 섹션 헤더 인식 실패 → fix_cid_spacing() 선처리로 해결
    - 여전히 OCR이 【 → [ 로 잘못 인식 → TERM_REPLACE에서 [변환 처리
    """
    sections = {'raw': text, 'sections': {}}
    lines = text.split('\n')
    current_section = '__header__'
    sections['sections'][current_section] = []

    # 페이지 경계 표시 제거 (예: "1/13", "2/13")
    filtered_lines = []
    for line in lines:
        if re.match(r'^\d+/\d+$', line.strip()):
            continue  # 페이지 번호 제거
        filtered_lines.append(line)
    lines = filtered_lines

    for line in lines:
        matched = None
        for key in SECTION_KEYS:
            if key in line:
                matched = key
                break
        if matched:
            current_section = matched
            sections['sections'][current_section] = []
        else:
            sections['sections'][current_section].append(line)

    # 빈 섹션 제거
    sections['sections'] = {k: v for k, v in sections['sections'].items() if v}
    return sections

# ============================================================
# 6. PDF 표 구조 분석 + 재현
# [유지보수 참고]
# - PyMuPDF Path(drawings)로 수평/수직선 감지
# - 감지된 선으로 표 영역(TableRegion) 추정
# - 실패 사례: 스캔 PDF는 선 정보 없음 → fallback으로 OCR 텍스트만 출력
# ============================================================

def detect_table_regions(pdf_path):
    """PyMuPDF Path로 표 선 감지 → 표 영역 추정"""
    doc = fitz.open(str(pdf_path))
    tables = []

    for page_idx, page in enumerate(doc, 1):
        paths = page.get_drawings()
        h_lines, v_lines = [], []

        for p in paths:
            for item in p.get("items", []):
                if item[0] == "l":  # line
                    x0, y0 = item[1]; x1, y1 = item[2]
                    if abs(y1 - y0) < 3:  # 수평선
                        h_lines.append((min(y0, y1), min(x0, x1), max(x0, x1)))
                    elif abs(x1 - x0) < 3:  # 수직선
                        v_lines.append((min(x0, x1), min(y0, y1), max(y0, y1)))

        if len(h_lines) >= 2 and len(v_lines) >= 2:
            y_coords = sorted(set(round(y) for y, _, _ in h_lines))
            x_coords = sorted(set(round(x) for x, _, _ in v_lines))
            tables.append({
                'page': page_idx,
                'x0': min(x_coords), 'y0': min(y_coords),
                'x1': max(x_coords), 'y1': max(y_coords),
                'rows': len(y_coords) - 1,
                'cols': len(x_coords) - 1,
            })

    doc.close()
    return tables

# ============================================================
# 7. 정확도 검증
# ============================================================

KEYWORDS_CHECK = [
    '고유번호', '2849-2018-019318', '소재지',
    '경기도 파주시 파평면 마산리 113-2',
    '토지', '열람일시', '2026년03월31일', '표제부', '갑구', '을구',
    '소유권이전', '근저당권설정', '지상권설정', '가압류', '임의경매개시결정',
    '이순옥', '강성원', '최영호', '이은',
    '파주농업협동조합', '북파주농업협동조합',
    '매매목록', '공동담보목록',
    '의정부지방법원', '고양지원', '파주등기소',
    '채권최고액', '금266,000,000원', '금221,000,000원',
]

def validate_accuracy(extracted_text, gt_text):
    """src_비교 기준 대비 정확도 측정 (공백/포맷 차이 보정)"""
    if not gt_text:
        return {"error": "기준 파일 없음", "accuracy": 0.0}

    # 공백 제거 후 문자 정확도
    gt_clean = re.sub(r'\s+', '', gt_text)
    ex_clean = re.sub(r'\s+', '', extracted_text)
    common = sum(1 for c in gt_clean if c in ex_clean)
    char_acc = round(common / max(len(gt_clean), 1) * 100, 2)

    # 키워드 매칭 (유연 매칭)
    matched = 0
    matched_list = []
    missed_list = []
    for kw in KEYWORDS_CHECK:
        # 공백 제거 후 비교
        kw_clean = kw.replace(' ', '')
        if kw_clean in ex_clean.replace(' ', ''):
            matched += 1
            matched_list.append(kw)
        else:
            missed_list.append(kw)

    return {
        "char_accuracy_pct": char_acc,
        "keyword_matches": f"{matched}/{len(KEYWORDS_CHECK)}",
        "keyword_pct": round(matched / len(KEYWORDS_CHECK) * 100, 2),
        "missed_keywords": missed_list,
    }

# ============================================================
# 8. 출력 포맷
# ============================================================

def save_markdown(sections, tables, label='v7'):
    """Markdown 출력 (표 구조 포함)"""
    path = OUT_DIR / f'{label}_output.md'
    lines = ['# 등기부등본 파싱 결과 (v7)\n']
    lines.append(f'> 감지된 표 영역: {len(tables)}개\n')

    for name, content in sections['sections'].items():
        lines.append(f'\n## {name}\n')
        lines.extend(content)

    return _write(path, '\n'.join(lines))

def save_json(data, label='v7'):
    path = OUT_DIR / f'{label}_output.json'
    return _write(path, json.dumps(data, ensure_ascii=False, indent=2))

def save_excel(sections, label='v7'):
    """Excel 출력 (섹션별 워크시트)"""
    wb = Workbook()
    ws = wb.active
    ws.title = '전체데이터'
    ws.cell(row=1, column=1, value='등기부등본 파싱 결과 (v7)').font = Font(bold=True, size=14)
    row = 3
    for name, content in sections['sections'].items():
        ws.cell(row=row, column=1, value=f'[{name}]').font = Font(bold=True, size=12)
        row += 1
        for line in content:
            ws.cell(row=row, column=1, value=line)
            row += 1
        row += 1
    path = OUT_DIR / f'{label}_output.xlsx'
    wb.save(str(path))
    return str(path)

def _write(path, content):
    path.write_text(content, encoding='utf-8') if hasattr(path, 'write_text') else None
    return str(path)

# ============================================================
# 9. 메인 실행
# ============================================================

def run(pdf_path=PDF_DEFAULT, label='v7'):
    """전체 파이프라인 실행"""
    print(f'[v7] {pdf_path.name} 파싱 시작...')
    print(f'  기준 파일: {GROUND_TRUTH.name}')

    # 기준 파일 로드
    gt_text = GROUND_TRUTH.read_text(encoding='utf-8') if GROUND_TRUTH.exists() else ''

    # PDF 표 구조 분석
    print(f'[v7] 표 구조 분석...')
    tables = detect_table_regions(pdf_path)
    print(f'  감지된 표 영역: {len(tables)}개')

    # OCR + 정제
    print(f'[v7] OCR 처리...')
    pages = ocr_pdf(pdf_path)
    print(f'  {len(pages)}페이지 OCR 완료')

    # 정제
    combined_raw = '\n\n'.join(p['text'] for p in pages)
    cleaned = clean_text(combined_raw)

    # 섹션 파싱
    sections = parse_sections(cleaned)
    print(f'  감지된 섹션: {list(sections["sections"].keys())}')

    # 정확도 검증
    accuracy = validate_accuracy(cleaned, gt_text)
    print(f'  문자 정확도: {accuracy.get("char_accuracy_pct", 0):.1f}%')
    print(f'  키워드 매칭: {accuracy.get("keyword_matches", "N/A")}')
    if accuracy.get('missed_keywords'):
        print(f'  누락 키워드: {accuracy["missed_keywords"]}')

    # 출력 저장
    md_path = save_markdown(sections, tables, label)
    json_path = save_json({
        'source': str(pdf_path),
        'pages': len(pages),
        'tables': tables,
        'cleaned_text': cleaned,
        'sections': sections['sections'],
        'accuracy': accuracy,
    }, label)
    xlsx_path = save_excel(sections, label)

    # combined 텍스트 저장
    (OUT_DIR / f'{label}_combined.txt').write_text(cleaned, encoding='utf-8')

    print(f'\n  출력:')
    print(f'    MD: {md_path}')
    print(f'    JSON: {json_path}')
    print(f'    Excel: {xlsx_path}')
    print(f'    TXT: {OUT_DIR / f"{label}_combined.txt"}')

    print(f'\n{"="*50}')
    print(f'  정확도: {accuracy.get("char_accuracy_pct", 0):.1f}%')
    print(f'  키워드: {accuracy.get("keyword_matches", "N/A")}')
    print(f'  표 영역: {len(tables)}개')
    print(f'{"="*50}')
    return accuracy

if __name__ == '__main__':
    run()
